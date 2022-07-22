from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from openpyxl import load_workbook
import pandas as pd

# definição das variaveis da string do exporta dados
status_eventos = 'inconsistencias'
data_inicio = '01/01/2022'
data_fim = '07/07/2022'

with open('empresa_mae.txt') as f_1:
    empresa_mae = f_1.read()

with open('chave.txt') as f_2:
    chave_ed = f_2.read()

# condição para status de evento 0 -> pendente; 2 -> inconsistencia
if status_eventos == 'inconsistencias':
    status = 2
elif status_eventos == 'pendentes':
    status = 0
else:
    print(f'status {status_eventos} não definido, escolha entre: inconsistencias ou pendentes')
    quit()

# definição dos caminhos para leitura e criação dos arquivos excel
leitura = r'empresas_esocial.xlsx'
criacao = rf'esocial_{status_eventos}_{data_inicio}_a_{data_fim}.xlsx'

# leitura do arquivo excel com nome e cod das empresas
empresas_df = pd.read_excel(leitura)

# lista de eventos para virar DF
evento = []

# condição para não inserir na lista caso só retorne o cabeçalho
condicao = 'EVENTO;DATAGERACAO;UNIDADE;NOMEUNIDADE;CODIGOGED;CODIGOARQUIVOGED;' \
            'NOMEARQUIVO;FUNCIONARIO;NOMEFUNCIONARIO;STATUSEVENTO;CODIGOLOTE;NRRECIBO;' \
            'IDARQUIVO;ERRO;SEQUENCIAL;AMBIENTEPRODUCAO;CODIGOERROESOCIAL;DATAINICIOCONDICAO;CARGAINICIAL'

# estrutura de repetição para os codigos das empresas do arquivo xlsx
for codigo in empresas_df['cod']:

    navegador = webdriver.Chrome()
    navegador.get("https://ws1.soc.com.br/WebSoc/exportadados?parametro={"
                  + f"'empresa':'{empresa_mae}','codigo':'141273',"
                    f"'chave':'{chave_ed}','tipoSaida':'txt','empresaTrabalho':{codigo}'"
                    f"','dataInicio':'{data_inicio}','dataFim':'{data_fim}','status':'{status}',"
                    "'layout':'0','unidade':'0','ambiente':'1','cabecalho':'1'}")
    csv = WebDriverWait(navegador, 50).until(ec.element_to_be_clickable((By.XPATH, '/html/body/pre'))).text

# tratativa do texto copiado do exporta dados gerando um xlsx para todas empresas
    plan = csv.split("\n")
    for linha in plan:
        if linha == condicao:
            continue
        coluna = linha.split(';')
        evento.append(coluna)

    navegador.close()

inconsistencias_df = pd.DataFrame(evento)

inconsistencias_df.to_excel(criacao, index=False)

# Alterando as planilhas
wb = load_workbook(criacao)
ws = wb.active

# removendo as colunas com dados desnecessarios
ws.delete_cols(19)
ws.delete_cols(16)
ws.delete_cols(15)
ws.delete_cols(13)
ws.delete_cols(12)
ws.delete_cols(11)
ws.delete_cols(10)
ws.delete_cols(7)
ws.delete_cols(6)
ws.delete_cols(5)
ws.delete_cols(3)

# dando nome para as colunas
ws['A1'] = 'EVENTO'
ws['B1'] = 'DATAGERACAO'
ws['C1'] = 'NOMEUNIDADE'
ws['D1'] = 'FUNCIONARIO'
ws['E1'] = 'NOMEFUNCIONARIO'
ws['F1'] = 'ERRO'
ws['G1'] = 'CODIGOERROESOCIAL'
ws['H1'] = 'DATAINICIOCONDICAO'

# alterando a largura das colunas
ws.column_dimensions['A'].width = 54
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 13
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 50
ws.column_dimensions['G'].width = 20
ws.column_dimensions['H'].width = 21

# salvando a planilha alterada
wb.save(rf'C:\Users\desenvolvimento02\Desktop\e-social {status_eventos} teste.xlsx')
